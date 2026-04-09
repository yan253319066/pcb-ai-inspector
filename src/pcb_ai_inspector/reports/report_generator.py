"""
PCB AI Inspector 报告生成模块。

从检测结果生成 PDF 和 Excel 报告。
"""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import Optional

from reportlab.platypus import Table, Paragraph, PageBreak, Spacer, TableStyle
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

from ..core.defect_types import DefectType, DEFECT_LABELS
from ..core.settings import ReportSettings
from ..ui.defect_overlay import DetectionResult


logger = logging.getLogger(__name__)


class ReportGenerator:
    """检测报告生成器，支持多种格式。"""

    def __init__(self, settings: Optional[ReportSettings] = None) -> None:
        """初始化报告生成器。

        Args:
            settings: 报告生成设置
        """
        self._settings = settings or ReportSettings()
        self._title = "PCB缺陷检测报告"
        self._company_name = self._settings.company_name or "PCB AI Inspector"

    def _setup_fonts(self, styles) -> tuple[str, str]:
        """设置字体。

        Args:
            styles: 样式表

        Returns:
            字体名称和粗体字体名称的元组
        """
        # 尝试设置中文字体
        try:
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont

            # 尝试注册中文字体
            try:
                # 尝试使用系统字体
                pdfmetrics.registerFont(TTFont("SimHei", "SimHei.ttf"))
                font_name = "SimHei"
                # 不使用 Bold 变体，避免字体不存在的问题
                bold_font_name = "SimHei"
            except Exception:
                # 如果系统没有中文字体，使用默认字体
                font_name = "Helvetica"
                bold_font_name = "Helvetica-Bold"
        except Exception:
            font_name = "Helvetica"
            bold_font_name = "Helvetica-Bold"
        return font_name, bold_font_name

    def _create_styles(self, styles, font_name) -> tuple:
        """创建样式。

        Args:
            styles: 样式表
            font_name: 字体名称

        Returns:
            标题样式、标题2样式、标题3样式的元组
        """
        from reportlab.lib.styles import ParagraphStyle

        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=18,
            alignment=1,  # Center
            spaceAfter=20,
            fontName=font_name,
        )
        heading_style = ParagraphStyle(
            "CustomHeading2",
            parent=styles["Heading2"],
            fontName=font_name,
        )
        subheading_style = ParagraphStyle(
            "CustomHeading3",
            parent=styles["Heading3"],
            fontName=font_name,
        )
        return title_style, heading_style, subheading_style

    def _create_metadata_table(self, meta_data, font_name, bold_font_name) -> Table:
        """创建元数据表。

        Args:
            meta_data: 元数据
            font_name: 字体名称
            bold_font_name: 粗体字体名称

        Returns:
            元数据表
        """
        from reportlab.platypus import Table, TableStyle
        from reportlab.lib.units import mm

        meta_table = Table(meta_data, colWidths=[30 * mm, 100 * mm])
        meta_table.setStyle(
            TableStyle(
                [
                    ("FONTNAME", (0, 0), (0, -1), bold_font_name),
                    ("FONTNAME", (1, 0), (-1, -1), font_name),
                    ("FONTSIZE", (0, 0), (-1, -1), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ]
            )
        )
        return meta_table

    def _create_summary_table(self, summary_data, font_name, bold_font_name) -> Table:
        """创建摘要表格。

        Args:
            summary_data: 摘要数据
            font_name: 字体名称
            bold_font_name: 粗体字体名称

        Returns:
            摘要表格
        """
        from reportlab.platypus import Table, TableStyle
        from reportlab.lib.units import mm

        summary_table = Table(summary_data, colWidths=[30 * mm, 50 * mm])
        summary_table.setStyle(
            TableStyle(
                [
                    ("FONTNAME", (0, 0), (0, -1), bold_font_name),
                    ("FONTNAME", (1, 0), (-1, -1), font_name),
                    ("FONTSIZE", (0, 0), (-1, -1), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )
        return summary_table

    def _create_detection_table(self, table_data, font_name, bold_font_name) -> Table:
        """创建缺陷详情表格。

        Args:
            table_data: 表格数据
            font_name: 字体名称
            bold_font_name: 粗体字体名称

        Returns:
            缺陷详情表格
        """
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import Table, TableStyle

        detection_table = Table(
            table_data,
            colWidths=[12 * mm, 25 * mm, 20 * mm, 30 * mm, 30 * mm, 30 * mm],
        )
        detection_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("FONTNAME", (0, 0), (-1, 0), bold_font_name),
                    ("FONTNAME", (0, 1), (-1, -1), font_name),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                    ("TOPPADDING", (0, 0), (-1, -1), 6),
                    (
                        "ROWBACKGROUNDS",
                        (0, 1),
                        (-1, -1),
                        [colors.white, colors.lightgrey],
                    ),
                ]
            )
        )
        return detection_table

    def _add_image(self, elements, image_path, heading_style) -> None:
        """添加图像到报告。

        Args:
            elements: 元素列表
            image_path: 图像路径
            heading_style: 标题样式
        """
        from reportlab.lib.units import mm
        from reportlab.platypus import Image, Spacer, Paragraph, PageBreak

        if image_path and image_path.exists():
            elements.append(PageBreak())  # 新起一页显示完整图像
            elements.append(Paragraph("原始图像", heading_style))
            try:
                # 先创建Image对象获取原始尺寸，然后缩放
                img = Image(str(image_path))

                # A4页面可用尺寸（减去边距和标题占用的空间）
                page_width = 160 * mm
                page_height = 230 * mm

                # img.imageWidth/imageHeight 返回的是原始图片像素尺寸
                # 需要转换为报告可用的尺寸（默认72 DPI）
                # ReportLab中 1 inch = 72 points, 图片像素 / 96 * 72 = points (假设96DPI原图)
                orig_width_pts = img.imageWidth * 72 / 96
                orig_height_pts = img.imageHeight * 72 / 96

                if orig_width_pts > page_width or orig_height_pts > page_height:
                    scale = min(
                        page_width / orig_width_pts, page_height / orig_height_pts
                    )
                    img.drawWidth = orig_width_pts * scale
                    img.drawHeight = orig_height_pts * scale
                else:
                    # 小图片使用原始尺寸（保持比例）
                    img.drawWidth = orig_width_pts
                    img.drawHeight = orig_height_pts

                img.hAlign = "CENTER"
                elements.append(img)
                elements.append(Spacer(1, 5 * mm))
            except Exception as e:
                logger.warning(f"Failed to include image: {e}")

    def _add_annotated_image(
        self,
        elements,
        image_path,
        detections,
        heading_style,
    ) -> None:
        """添加标注图像到报告。

        Args:
            elements: 元素列表
            image_path: 图像路径
            detections: 检测结果列表
            heading_style: 标题样式
        """
        from reportlab.lib.units import mm
        from reportlab.platypus import Image, Spacer, Paragraph, PageBreak
        import cv2
        import io
        from ..ui.defect_overlay import draw_detections_on_image

        if image_path and image_path.exists() and detections:
            elements.append(PageBreak())  # 新起一页显示完整图像
            elements.append(Paragraph("标注图像", heading_style))
            try:
                # 读取图像
                image = cv2.imread(str(image_path))
                if image is not None:
                    # 绘制检测结果
                    annotated_image = draw_detections_on_image(image, detections)
                    # 使用 BytesIO 缓冲区，避免文件系统的时序问题
                    buffer = io.BytesIO()
                    # 使用 PNG 格式（比 JPEG 更可靠）
                    success, encoded = cv2.imencode(".png", annotated_image)
                    if success:
                        buffer.write(encoded)
                        buffer.seek(0)
                        try:
                            # 从缓冲区读取图片数据
                            img = Image(buffer)
                            # A4页面可用尺寸（减去边距和标题占用的空间）
                            page_width = 160 * mm
                            page_height = 230 * mm

                            # 获取原始图片尺寸（像素）并转换为报告单位
                            # ReportLab 中 1 inch = 72 points, 图片像素 / 96 * 72 = points (假设96DPI)
                            orig_width_pts = img.imageWidth * 72 / 96
                            orig_height_pts = img.imageHeight * 72 / 96

                            if (
                                orig_width_pts > page_width
                                or orig_height_pts > page_height
                            ):
                                scale = min(
                                    page_width / orig_width_pts,
                                    page_height / orig_height_pts,
                                )
                                img.drawWidth = orig_width_pts * scale
                                img.drawHeight = orig_height_pts * scale
                            else:
                                img.drawWidth = orig_width_pts
                                img.drawHeight = orig_height_pts

                            img.hAlign = "CENTER"
                            elements.append(img)
                            elements.append(Spacer(1, 5 * mm))
                        except Exception as e:
                            logger.warning(
                                f"Failed to add annotated image to report: {e}"
                            )
                    else:
                        logger.warning("Failed to encode annotated image")
            except Exception as e:
                logger.warning(f"Failed to include annotated image: {e}")

    def generate_pdf(
        self,
        image_path: Optional[Path],
        detections: list[DetectionResult],
        output_path: Path,
        include_image: Optional[bool] = None,
    ) -> None:
        """生成 PDF 报告。

        Args:
            image_path: 源图像路径
            detections: 检测结果列表
            output_path: 输出文件路径
            include_image: 是否在报告中包含图像（None 表示使用设置中的值）
        """
        # 使用设置中的值或默认值
        include_image = (
            include_image if include_image is not None else self._settings.include_image
        )
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import mm
            from reportlab.platypus import (
                SimpleDocTemplate,
                Paragraph,
                Spacer,
            )
        except ImportError:
            logger.error(
                "reportlab is not installed. Please install it to generate PDF reports."
            )
            raise ImportError(
                "reportlab is required for PDF generation. Run: pip install reportlab"
            )

        # 创建文档
        doc = SimpleDocTemplate(
            str(output_path),
            pagesize=A4,
            rightMargin=20 * mm,
            leftMargin=20 * mm,
            topMargin=20 * mm,
            bottomMargin=20 * mm,
        )

        # 样式
        styles = getSampleStyleSheet()

        # 设置字体和样式
        font_name, bold_font_name = self._setup_fonts(styles)
        title_style, heading_style, _ = self._create_styles(styles, font_name)

        # Build content
        elements = []

        # Title
        elements.append(Paragraph(self._title, title_style))
        elements.append(Spacer(1, 10 * mm))

        # Metadata
        meta_data = [
            ["报告时间:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["检测设备:", "PCB AI Inspector"],
            ["图像文件:", image_path.name if image_path else "N/A"],
        ]

        # 添加公司名称（如果有）
        if self._settings.company_name:
            meta_data.insert(0, ["公司名称:", self._settings.company_name])
        meta_table = self._create_metadata_table(meta_data, font_name, bold_font_name)
        elements.append(meta_table)
        elements.append(Spacer(1, 10 * mm))

        # 添加原始图像（如果有）
        if include_image:
            self._add_image(elements, image_path, heading_style)
            # 添加标注图像（如果有检测结果）
            if detections:
                self._add_annotated_image(
                    elements, image_path, detections, heading_style
                )

        # Summary
        elements.append(Paragraph("检测摘要", heading_style))

        total_defects = len(detections)
        summary_data = [
            ["总缺陷数:", str(total_defects)],
        ]

        # 计算额外统计信息
        if detections:
            # 平均置信度
            avg_conf = sum(d.confidence for d in detections) / len(detections)
            summary_data.append(["平均置信度:", f"{avg_conf:.1%}"])

            # 最高置信度
            max_conf = max(d.confidence for d in detections)
            summary_data.append(["最高置信度:", f"{max_conf:.1%}"])

            # 最低置信度
            min_conf = min(d.confidence for d in detections)
            summary_data.append(["最低置信度:", f"{min_conf:.1%}"])

        # Count by type
        type_counts: dict[DefectType, int] = {dt: 0 for dt in DefectType}
        for d in detections:
            type_counts[d.defect_type] += 1

        for dt, count in type_counts.items():
            if count > 0:
                label = DEFECT_LABELS.get(dt.value, dt.value)
                percentage = (count / total_defects) * 100 if total_defects > 0 else 0
                summary_data.append([f"{label}:", f"{count} ({percentage:.1%})"])

        summary_table = self._create_summary_table(
            summary_data, font_name, bold_font_name
        )
        elements.append(summary_table)
        elements.append(Spacer(1, 10 * mm))

        # Detection details
        elements.append(Paragraph("缺陷详情", heading_style))
        if detections:
            # Table header
            table_data = [
                ["序号", "类型", "置信度", "位置 (x1, y1)", "位置 (x2, y2)", "尺寸"]
            ]

            for idx, det in enumerate(detections, 1):
                label = DEFECT_LABELS.get(det.defect_type.value, det.defect_type.value)
                table_data.append(
                    [
                        str(idx),
                        label,
                        f"{det.confidence:.1%}",
                        f"({det.x1:.0f}, {det.y1:.0f})",
                        f"({det.x2:.0f}, {det.y2:.0f})",
                        f"{det.width:.0f} x {det.height:.0f}",
                    ]
                )

            # Create table
            detection_table = self._create_detection_table(
                table_data, font_name, bold_font_name
            )
            elements.append(detection_table)
        else:
            # 创建无缺陷提示，使用正确的中文字体
            from reportlab.lib.styles import ParagraphStyle
            from reportlab.platypus import Paragraph

            no_defect_style = ParagraphStyle(
                "NoDefectStyle",
                fontName=font_name,
                fontSize=10,
                leading=12,
            )
            elements.append(Paragraph("无缺陷", no_defect_style))

        # Build PDF
        doc.build(elements)
        logger.info(f"PDF report generated: {output_path}")

    def _add_excel_images(
        self,
        ws,
        image_path: Optional[Path],
        detections: list[DetectionResult],
        start_row: int,
    ) -> int:
        """在Excel工作表中添加图片。

        Args:
            ws: 工作表
            image_path: 图片路径
            detections: 检测结果列表
            start_row: 起始行

        Returns:
            下一个可用的行号
        """
        from openpyxl.drawing.image import Image as XLImage
        import cv2
        import io
        from ..ui.defect_overlay import draw_detections_on_image

        if not (image_path and image_path.exists()):
            return start_row

        current_row = start_row

        # 设置图片大小
        img_width = 280
        img_height = 210

        # 添加标题行
        ws.cell(row=current_row, column=1, value="原始图像")
        if detections:
            ws.cell(row=current_row, column=5, value="标注图像")
        current_row += 1

        # 设置图片区域的行高（确保图片有足够空间）
        for r in range(current_row, current_row + 12):
            ws.row_dimensions[r].height = 15

        # 原始图像放在 A 列
        try:
            img = XLImage(str(image_path))
            img.width = img_width
            img.height = img_height
            ws.add_image(img, f"A{current_row}")
        except Exception as e:
            logger.warning(f"Failed to add original image to Excel: {e}")

        # 标注图像放在 E 列（与原图并排）
        if detections:
            try:
                image = cv2.imread(str(image_path))
                if image is not None:
                    annotated_image = draw_detections_on_image(image, detections)
                    buffer = io.BytesIO()
                    success, encoded = cv2.imencode(".png", annotated_image)
                    if success:
                        buffer.write(encoded)
                        buffer.seek(0)
                        img = XLImage(buffer)
                        img.width = img_width
                        img.height = img_height
                        ws.add_image(img, f"E{current_row}")
            except Exception as e:
                logger.warning(f"Failed to add annotated image to Excel: {e}")

        # 返回图片后面的空行（留足够空间给图片和数据之间的间隔）
        return current_row + 14

    def generate_excel(
        self,
        detections: list[DetectionResult],
        output_path: Path,
        image_path: Optional[Path] = None,
        include_image: Optional[bool] = None,
    ) -> None:
        """生成 Excel 报告。

        Args:
            detections: 检测结果列表
            output_path: 输出文件路径
            image_path: 源图像路径
            include_image: 是否包含图像（None 表示使用设置中的值）
        """
        # 使用设置中的值或默认值
        include_image = (
            include_image if include_image is not None else self._settings.include_image
        )

        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            logger.error(
                "openpyxl is not installed. Please install it to generate Excel reports."
            )
            raise ImportError(
                "openpyxl is required for Excel generation. Run: pip install openpyxl"
            )

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            raise RuntimeError("Failed to create worksheet")

        ws.title = "检测报告"

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Title
        ws.merge_cells("A1:G1")
        ws["A1"] = self._title
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = center_align

        # Metadata
        row = 3
        ws[f"A{row}"] = "报告时间:"
        ws[f"B{row}"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        row += 1

        # 添加公司名称（如果有）
        if self._settings.company_name:
            ws[f"A{row}"] = "公司名称:"
            ws[f"B{row}"] = self._settings.company_name
            row += 1

        ws[f"A{row}"] = "总缺陷数:"
        ws[f"B{row}"] = len(detections)
        row += 1

        # 添加额外统计信息
        if detections:
            avg_conf = sum(d.confidence for d in detections) / len(detections)
            max_conf = max(d.confidence for d in detections)
            min_conf = min(d.confidence for d in detections)

            ws[f"A{row}"] = "平均置信度:"
            ws[f"B{row}"] = avg_conf
            ws[f"B{row}"].number_format = "0.0%"
            row += 1

            ws[f"A{row}"] = "最高置信度:"
            ws[f"B{row}"] = max_conf
            ws[f"B{row}"].number_format = "0.0%"
            row += 1

            ws[f"A{row}"] = "最低置信度:"
            ws[f"B{row}"] = min_conf
            ws[f"B{row}"].number_format = "0.0%"
            row += 1

            # 按类型统计
            type_counts: dict[str, int] = {}
            for det in detections:
                label = DEFECT_LABELS.get(det.defect_type.value, det.defect_type.value)
                if label not in type_counts:
                    type_counts[label] = 0
                type_counts[label] += 1

            ws[f"A{row}"] = "缺陷类型分布:"
            row += 1
            for label, count in type_counts.items():
                percentage = (count / len(detections)) * 100
                ws[f"A{row}"] = f"  {label}:"
                ws[f"B{row}"] = f"{count} ({percentage:.1f}%)"
                row += 1

        # 添加图片（如果有）
        if include_image:
            row += 1
            row = self._add_excel_images(ws, image_path, detections, row)

        # Table header
        table_header_row = row + 1
        headers = ["序号", "缺陷类型", "置信度", "X1", "Y1", "X2", "Y2", "宽度", "高度"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=table_header_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        # Data rows
        for row_idx, det in enumerate(detections, 1):
            label = DEFECT_LABELS.get(det.defect_type.value, det.defect_type.value)
            row = row_idx + table_header_row

            data = [
                row_idx,
                label,
                det.confidence,
                det.x1,
                det.y1,
                det.x2,
                det.y2,
                det.width,
                det.height,
            ]

            for col_idx, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.alignment = center_align
                cell.border = thin_border

                # Format confidence as percentage
                if col_idx == 3:
                    cell.number_format = "0.0%"

        # Adjust column widths
        column_widths = [8, 15, 12, 10, 10, 10, 10, 10, 10]
        for col_idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col_idx)
            ].width = width

        # Save
        wb.save(str(output_path))
        logger.info(f"Excel report generated: {output_path}")

    def generate_batch_pdf(
        self,
        results: dict[str, dict],
        output_path: Path,
        include_image: Optional[bool] = None,
    ) -> None:
        """生成批量 PDF 报告。

        Args:
            results: 批量检测结果字典
            output_path: 输出文件路径
            include_image: 是否在报告中包含图像（None 表示使用设置中的值）
        """
        # 使用设置中的值或默认值
        include_image = (
            include_image if include_image is not None else self._settings.include_image
        )
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import mm
            from reportlab.platypus import (
                SimpleDocTemplate,
                Paragraph,
                Spacer,
            )
        except ImportError:
            logger.error(
                "reportlab is not installed. Please install it to generate PDF reports."
            )
            raise ImportError(
                "reportlab is required for PDF generation. Run: pip install reportlab"
            )

        # 创建文档
        doc = SimpleDocTemplate(
            str(output_path),
            pagesize=A4,
            rightMargin=20 * mm,
            leftMargin=20 * mm,
            topMargin=20 * mm,
            bottomMargin=20 * mm,
        )

        # 样式
        styles = getSampleStyleSheet()

        # 设置字体和样式
        font_name, bold_font_name = self._setup_fonts(styles)
        title_style, heading_style, subheading_style = self._create_styles(
            styles, font_name
        )

        # Build content
        elements = []

        # Title
        elements.append(Paragraph(self._title, title_style))
        elements.append(Spacer(1, 10 * mm))

        # Metadata
        meta_data = [
            ["报告时间:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["检测设备:", "PCB AI Inspector"],
            ["检测图像数:", str(len(results))],
        ]

        # 添加公司名称（如果有）
        if self._settings.company_name:
            meta_data.insert(0, ["公司名称:", self._settings.company_name])
        meta_table = self._create_metadata_table(meta_data, font_name, bold_font_name)
        elements.append(meta_table)
        elements.append(Spacer(1, 10 * mm))

        # 总体统计
        elements.append(Paragraph("总体统计", heading_style))
        total_defects = 0
        total_confidence = 0
        defect_count = 0
        type_counts: dict[str, int] = {}

        for result in results.values():
            detections = result.get("detections", [])
            total_defects += len(detections)
            for det in detections:
                total_confidence += det.confidence
                defect_count += 1
                label = DEFECT_LABELS.get(det.defect_type.value, det.defect_type.value)
                if label not in type_counts:
                    type_counts[label] = 0
                type_counts[label] += 1

        summary_data = [
            ["总缺陷数:", str(total_defects)],
            [
                "平均每图缺陷数:",
                f"{total_defects / len(results):.2f}" if results else "0",
            ],
        ]

        # 添加额外统计信息
        if defect_count > 0:
            avg_conf = total_confidence / defect_count
            summary_data.append(["平均置信度:", f"{avg_conf:.1%}"])

        for label, count in type_counts.items():
            percentage = (count / total_defects) * 100 if total_defects > 0 else 0
            summary_data.append([f"{label}:", f"{count} ({percentage:.1%})"])

        summary_table = self._create_summary_table(
            summary_data, font_name, bold_font_name
        )
        elements.append(summary_table)
        elements.append(Spacer(1, 15 * mm))

        # 按图像逐个生成报告
        for image_path_str, result in results.items():
            image_path = Path(image_path_str)
            detections = result.get("detections", [])

            # 图像标题
            elements.append(Paragraph(f"图像: {image_path.name}", subheading_style))
            elements.append(Spacer(1, 5 * mm))

            # 显示图像（如果有）
            if include_image:
                self._add_image(elements, image_path, subheading_style)
                # 添加标注图像（如果有检测结果）
                if detections:
                    self._add_annotated_image(
                        elements,
                        image_path,
                        detections,
                        subheading_style,
                    )

            # 图像级统计
            image_summary = [
                ["缺陷数:", str(len(detections))],
            ]
            if detections:
                avg_conf = sum(d.confidence for d in detections) / len(detections)
                max_conf = max(d.confidence for d in detections)
                min_conf = min(d.confidence for d in detections)
                image_summary.append(["平均置信度:", f"{avg_conf:.1%}"])
                image_summary.append(["最高置信度:", f"{max_conf:.1%}"])
                image_summary.append(["最低置信度:", f"{min_conf:.1%}"])

            image_summary_table = self._create_summary_table(
                image_summary, font_name, bold_font_name
            )
            elements.append(image_summary_table)
            elements.append(Spacer(1, 5 * mm))

            # 缺陷详情
            if detections:
                # Table header
                table_data = [
                    ["序号", "类型", "置信度", "位置 (x1, y1)", "位置 (x2, y2)", "尺寸"]
                ]

                for idx, det in enumerate(detections, 1):
                    label = DEFECT_LABELS.get(
                        det.defect_type.value, det.defect_type.value
                    )
                    table_data.append(
                        [
                            str(idx),
                            label,
                            f"{det.confidence:.1%}",
                            f"({det.x1:.0f}, {det.y1:.0f})",
                            f"({det.x2:.0f}, {det.y2:.0f})",
                            f"{det.width:.0f} x {det.height:.0f}",
                        ]
                    )

                # Create table
                detection_table = self._create_detection_table(
                    table_data, font_name, bold_font_name
                )
                elements.append(detection_table)
            else:
                # 创建无缺陷提示，使用正确的中文字体
                no_defect_style = ParagraphStyle(
                    "NoDefectStyle",
                    fontName=font_name,
                    fontSize=10,
                    leading=12,
                )
                elements.append(Paragraph("无缺陷", no_defect_style))

            elements.append(Spacer(1, 15 * mm))

        # 添加附录（原始大小标注图像）- 如果有图片和检测结果
        # Build PDF
        doc.build(elements)
        logger.info(f"Batch PDF report generated: {output_path}")

    def generate_batch_excel(
        self,
        results: dict[str, dict],
        output_path: Path,
        include_image: Optional[bool] = None,
    ) -> None:
        """生成批量 Excel 报告。

        Args:
            results: 批量检测结果字典
            output_path: 输出文件路径
            include_image: 是否包含图像（None 表示使用设置中的值）
        """
        # 使用设置中的值或默认值
        include_image = (
            include_image if include_image is not None else self._settings.include_image
        )

        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            logger.error(
                "openpyxl is not installed. Please install it to generate Excel reports."
            )
            raise ImportError(
                "openpyxl is required for Excel generation. Run: pip install openpyxl"
            )

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            raise RuntimeError("Failed to create worksheet")

        ws.title = "批量检测报告"

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        subheader_font = Font(bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Title
        ws.merge_cells("A1:I1")
        ws["A1"] = self._title
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = center_align

        # Metadata
        row = 3
        ws[f"A{row}"] = "报告时间:"
        ws[f"B{row}"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        row += 1

        # 添加公司名称（如果有）
        if self._settings.company_name:
            ws[f"A{row}"] = "公司名称:"
            ws[f"B{row}"] = self._settings.company_name
            row += 1

        ws[f"A{row}"] = "检测图像数:"
        ws[f"B{row}"] = len(results)
        row += 1

        # 总体统计
        total_defects = 0
        total_confidence = 0
        defect_count = 0
        type_counts: dict[str, int] = {}

        for result in results.values():
            detections = result.get("detections", [])
            total_defects += len(detections)
            for det in detections:
                total_confidence += det.confidence
                defect_count += 1
                label = DEFECT_LABELS.get(det.defect_type.value, det.defect_type.value)
                if label not in type_counts:
                    type_counts[label] = 0
                type_counts[label] += 1

        ws[f"A{row}"] = "总缺陷数:"
        ws[f"B{row}"] = total_defects
        row += 1
        ws[f"A{row}"] = "平均每图缺陷数:"
        ws[f"B{row}"] = total_defects / len(results) if results else 0
        row += 1

        # 添加额外统计信息
        if defect_count > 0:
            avg_conf = total_confidence / defect_count
            ws[f"A{row}"] = "平均置信度:"
            ws[f"B{row}"] = avg_conf
            ws[f"B{row}"].number_format = "0.0%"
            row += 1

        # 按类型统计
        ws[f"A{row}"] = "缺陷类型分布:"
        row += 1
        for label, count in type_counts.items():
            percentage = (count / total_defects) * 100 if total_defects > 0 else 0
            ws[f"A{row}"] = f"  {label}:"
            ws[f"B{row}"] = f"{count} ({percentage:.1f}%)"
            row += 1

        # Table header
        table_header_row = row + 2
        ws[f"A{table_header_row}"] = "图像文件名"
        ws[f"B{table_header_row}"] = "序号"
        ws[f"C{table_header_row}"] = "缺陷类型"
        ws[f"D{table_header_row}"] = "置信度"
        ws[f"E{table_header_row}"] = "X1"
        ws[f"F{table_header_row}"] = "Y1"
        ws[f"G{table_header_row}"] = "X2"
        ws[f"H{table_header_row}"] = "Y2"
        ws[f"I{table_header_row}"] = "尺寸"

        # Apply header styles
        for col in range(1, 10):
            cell = ws.cell(row=table_header_row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        # Data rows
        row = table_header_row + 1
        for image_path_str, result in results.items():
            image_path = Path(image_path_str)
            detections = result.get("detections", [])

            # Image filename row
            ws.merge_cells(f"A{row}:I{row}")
            cell = ws.cell(row=row, column=1, value=image_path.name)
            cell.font = subheader_font
            row += 1

            # 添加图片（如果有）
            if include_image:
                row = self._add_excel_images(ws, image_path, detections, row)

            # Detection rows
            for idx, det in enumerate(detections, 1):
                label = DEFECT_LABELS.get(det.defect_type.value, det.defect_type.value)
                size = f"{det.width:.0f} x {det.height:.0f}"

                data = [
                    "",  # Image filename (empty for data rows)
                    idx,
                    label,
                    det.confidence,
                    det.x1,
                    det.y1,
                    det.x2,
                    det.y2,
                    size,
                ]

                for col_idx, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col_idx, value=value)
                    cell.alignment = center_align
                    cell.border = thin_border

                    # Format confidence as percentage
                    if col_idx == 4:
                        cell.number_format = "0.0%"

                row += 1

            # Add empty row between images
            row += 1

        # Adjust column widths
        column_widths = [30, 8, 15, 12, 10, 10, 10, 10, 15]
        for col_idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col_idx)
            ].width = width

        # Save
        wb.save(str(output_path))
        logger.info(f"Batch Excel report generated: {output_path}")

    def generate_shift_report(
        self,
        records: list,
        shift_name: str,
        output_path: Path,
    ) -> None:
        """生成班次报表（Excel）。

        Args:
            records: 检测记录列表 (DetectionRecord)
            shift_name: 班次名称
            output_path: 输出文件路径
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            raise ImportError("openpyxl is required for Excel generation")

        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            raise RuntimeError("Failed to create worksheet")
        ws.title = f"{shift_name}班次报表"

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="22C55E", end_color="22C55E", fill_type="solid"
        )
        title_font = Font(bold=True, size=14)
        center_align = Alignment(horizontal="center", vertical="center")

        # Title
        ws.merge_cells("A1:G1")
        ws["A1"] = f"班次检测报表 - {shift_name}"
        ws["A1"].font = title_font
        ws["A1"].alignment = center_align

        # Statistics
        total = len(records)
        passed = sum(1 for r in records if r.result == "PASS")
        failed = total - passed
        pass_rate = (passed / total * 100) if total > 0 else 100

        ws["A3"] = "统计数据"
        ws["A3"].font = Font(bold=True)
        ws["A4"] = f"检测总数: {total}"
        ws["A5"] = f"良品数: {passed}"
        ws["A6"] = f"不良品数: {failed}"
        ws["A7"] = f"良率: {pass_rate:.1f}%"

        # Headers
        headers = ["时间", "生产线", "工位", "班次", "结果", "缺陷数", "处理时间(ms)"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=9, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        # Data
        for row_idx, record in enumerate(records, 10):
            ws.cell(row=row_idx, column=1, value=str(record.timestamp)[:19])
            ws.cell(row=row_idx, column=2, value=record.production_line or "")
            ws.cell(row=row_idx, column=3, value=record.station_name or "")
            ws.cell(row=row_idx, column=4, value=record.shift_config or "")
            ws.cell(row=row_idx, column=5, value=record.result)
            ws.cell(row=row_idx, column=6, value=record.defect_count)
            ws.cell(row=row_idx, column=7, value=record.processing_time_ms or 0)

        # Column widths
        for col_idx, width in enumerate([20, 15, 12, 12, 10, 10, 15], 1):
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col_idx)
            ].width = width

        wb.save(str(output_path))
        logger.info(f"Shift report generated: {output_path}")

    def generate_daily_report(
        self,
        records: list,
        date: str,
        output_path: Path,
    ) -> None:
        """生成日报表（Excel）。

        Args:
            records: 检测记录列表 (DetectionRecord)
            date: 日期字符串 YYYY-MM-DD
            output_path: 输出文件路径
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            raise ImportError("openpyxl is required for Excel generation")

        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            raise RuntimeError("Failed to create worksheet")
        ws.title = f"{date}日报表"

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="3B82F6", end_color="3B82F6", fill_type="solid"
        )
        title_font = Font(bold=True, size=14)
        center_align = Alignment(horizontal="center", vertical="center")

        # Title
        ws.merge_cells("A1:G1")
        ws["A1"] = f"日报表 - {date}"
        ws["A1"].font = title_font
        ws["A1"].alignment = center_align

        # Statistics
        total = len(records)
        passed = sum(1 for r in records if r.result == "PASS")
        failed = total - passed
        pass_rate = (passed / total * 100) if total > 0 else 100

        # Defect statistics
        defect_counts: dict[str, int] = {}
        for r in records:
            if r.defect_count > 0:
                defect_counts["总计"] = defect_counts.get("总计", 0) + r.defect_count

        ws["A3"] = "统计数据"
        ws["A3"].font = Font(bold=True)
        ws["A4"] = f"检测总数: {total}"
        ws["A5"] = f"良品数: {passed}"
        ws["A6"] = f"不良品数: {failed}"
        ws["A7"] = f"良率: {pass_rate:.1f}%"

        # Headers
        headers = ["时间", "生产线", "工位", "班次", "结果", "缺陷数", "处理时间(ms)"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=9, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        # Data
        for row_idx, record in enumerate(records, 10):
            ws.cell(row=row_idx, column=1, value=str(record.timestamp)[:19])
            ws.cell(row=row_idx, column=2, value=record.production_line or "")
            ws.cell(row=row_idx, column=3, value=record.station_name or "")
            ws.cell(row=row_idx, column=4, value=record.shift_config or "")
            ws.cell(row=row_idx, column=5, value=record.result)
            ws.cell(row=row_idx, column=6, value=record.defect_count)
            ws.cell(row=row_idx, column=7, value=record.processing_time_ms or 0)

        # Column widths
        for col_idx, width in enumerate([20, 15, 12, 12, 10, 10, 15], 1):
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col_idx)
            ].width = width

        wb.save(str(output_path))
        logger.info(f"Daily report generated: {output_path}")
